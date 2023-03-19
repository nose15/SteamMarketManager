from src.steamAPIwrapper.steamAPI import get_user_inventory, get_inventory_prices
import openpyxl
import asyncio


EXCEL_PATH = "skins.xlsx"


async def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb.active

    app_id = 730    # CS:GO
    currency_code = 6   # PLN

    inventory = get_user_inventory("76561198151374664", app_id)

    for idx, key in enumerate(inventory):
        item = inventory[key]
        r = (idx+2)
        print(item)
        sheet.cell(r, column=1).value = item['name']
        sheet.cell(r, column=2).value = item['amount']

    prices = await get_inventory_prices(app_id, currency_code, inventory)

    for idx, item_price in enumerate(prices):
        price = prices[item_price]
        r = idx + 2
        sheet.cell(r, column=4).value = price

    wb.save(EXCEL_PATH)
    wb.close()


if __name__ == '__main__':
    asyncio.run(main())

